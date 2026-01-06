# SPDX-License-Identifier: MIT
"""Scopus journal list data source (optional user-provided Excel file)."""

import asyncio
import glob
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ...cache import DataSourceManager
from ...config import get_config_manager
from ...enums import AssessmentType
from ...logging_config import get_detail_logger, get_status_logger
from ...normalizer import input_normalizer
from ...validation import validate_issn
from ..core import DataSource


detail_logger = get_detail_logger()
status_logger = get_status_logger()


class ScopusSource(DataSource):
    """Data source for Scopus journal list (optional user-provided Excel file)."""

    def __init__(self, data_dir: Path | None = None) -> None:
        """Initialize Scopus source.

        Args:
            data_dir: Directory to search for Scopus files.
                     Defaults to .aletheia-probe/scopus/ in current directory
        """
        if data_dir is None:
            data_dir = Path.cwd() / ".aletheia-probe" / "scopus"
        self.data_dir = data_dir
        self.file_path: Path | None = None

        # Load column mappings from configuration
        config = get_config_manager().load_config()
        self.column_mappings = config.data_source_processing.scopus_column_mappings

    def get_name(self) -> str:
        return "scopus"

    def get_list_type(self) -> AssessmentType:
        return AssessmentType.LEGITIMATE

    def should_update(self) -> bool:
        """Check if we should update (monthly for static file)."""
        # First check if file exists
        if not self._find_scopus_file():
            self.skip_reason = "file_not_found"
            return False

        data_source_manager = DataSourceManager()
        last_update = data_source_manager.get_source_last_updated(self.get_name())
        if last_update is None:
            return True

        # Update monthly
        if (datetime.now() - last_update).days < 30:
            self.skip_reason = "already_up_to_date"
            return False

        return True

    def _find_scopus_file(self) -> bool:
        """Find the most recent Scopus Excel file in the data directory.

        Returns:
            True if file found, False otherwise
        """
        if not self.data_dir.exists():
            self.data_dir.mkdir(parents=True, exist_ok=True)

        # Look for ext_list_*.xlsx files
        pattern = str(self.data_dir / "ext_list_*.xlsx")
        matching_files = glob.glob(pattern)

        if not matching_files:
            status_logger.info(
                f"    {self.get_name()}: No Scopus journal list found in {self.data_dir}"
            )
            detail_logger.info(
                f"No Scopus journal list found in {self.data_dir}. "
                "To use Scopus data, download the latest list from "
                '"https://www.researchgate.net/publication/384898389_Last_Update_of_Scopus_Indexed_Journal\'s_List_-_October_2024" '
                "and place it in this directory."
            )
            return False

        # Use the most recent file
        self.file_path = Path(
            max(matching_files, key=lambda p: Path(p).stat().st_mtime)
        )
        status_logger.info(
            f"    {self.get_name()}: Found journal list: {self.file_path.name}"
        )
        return True

    def _find_column_indices(self, headers: list[Any]) -> dict[str, int]:
        """Find column indices using configured mappings.

        Args:
            headers: List of header values from the Excel sheet

        Returns:
            Dictionary mapping field names to column indices
        """
        col_indices = {}
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                # Check each configured mapping
                for field_name, possible_headers in self.column_mappings.items():
                    # Skip if we've already found this field
                    if field_name in col_indices:
                        continue

                    if field_name == "quality_flag":
                        # Special case: quality_flag needs both "discontinued" AND "quality"
                        if all(keyword in header_lower for keyword in possible_headers):
                            col_indices[field_name] = i
                    else:
                        # Regular case: match any of the possible headers
                        for possible_header in possible_headers:
                            if possible_header in header_lower:
                                col_indices[field_name] = i
                                break
        return col_indices

    def _validate_and_normalize_issn(
        self, issn: str | None, journal_title: str
    ) -> str | None:
        """Validate and normalize ISSN format.

        Args:
            issn: Raw ISSN value
            journal_title: Journal title for logging

        Returns:
            Normalized ISSN or None if invalid
        """
        if not issn:
            return None

        issn = str(issn).strip()
        # Ensure hyphen format (NNNN-NNNN)
        if len(issn) == 8 and "-" not in issn:
            issn = f"{issn[:4]}-{issn[4:]}"

        # Validate ISSN checksum
        if not validate_issn(issn):
            detail_logger.warning(
                f"Invalid ISSN '{issn}' for journal '{journal_title}' - skipping ISSN"
            )
            return None

        return issn

    def _extract_cell_value(
        self, row: tuple[Any, ...], col_indices: dict[str, int], field_name: str
    ) -> Any:
        """Extract cell value from row using column indices.

        Args:
            row: Excel row tuple
            col_indices: Dictionary mapping field names to column indices
            field_name: Name of the field to extract

        Returns:
            Cell value or None if field not in col_indices
        """
        return row[col_indices[field_name]].value if field_name in col_indices else None

    def _create_journal_entry(
        self,
        title: str,
        issn: str | None,
        eissn: str | None,
        publisher: Any,
        source_type: Any,
        coverage: Any,
        open_access: Any,
        is_quality_flagged: bool,
        quality_flag: Any,
    ) -> dict[str, Any] | None:
        """Create journal entry from row data.

        Args:
            title: Journal title
            issn: Normalized ISSN
            eissn: Normalized e-ISSN
            publisher: Publisher name
            source_type: Source type value
            coverage: Coverage information
            open_access: Open access information
            is_quality_flagged: Whether journal has quality flag
            quality_flag: Quality flag value

        Returns:
            Journal entry dictionary or None if normalization fails
        """
        try:
            normalized_input = input_normalizer.normalize(title)

            metadata: dict[str, Any] = {
                "source_type": (str(source_type).strip() if source_type else "Journal"),
                "coverage": str(coverage).strip() if coverage else None,
                "open_access": str(open_access).strip() if open_access else None,
            }

            # Add quality flag if present
            if is_quality_flagged:
                metadata["quality_flagged"] = True
                metadata["quality_flag_reason"] = (
                    str(quality_flag).strip() if quality_flag else None
                )

            return {
                "journal_name": title,
                "normalized_name": normalized_input.normalized_name,
                "issn": issn,
                "eissn": eissn,
                "publisher": str(publisher).strip() if publisher else None,
                "metadata": metadata,
            }

        except Exception as e:
            detail_logger.debug(f"Failed to process journal '{title}': {e}")
            return None

    async def fetch_data(self) -> list[dict[str, Any]]:
        """Fetch and parse Scopus journal data from Excel file."""
        if not self._find_scopus_file():
            return []

        if not self.file_path:
            return []

        status_logger.info(
            f"    {self.get_name()}: Loading journal list from {self.file_path.name}"
        )

        try:
            # Load workbook in thread pool to avoid blocking event loop
            workbook = await asyncio.to_thread(
                load_workbook, filename=self.file_path, read_only=True, data_only=True
            )

            # Find the sheet (usually "Scopus Sources Oct. 2024" or similar)
            sheet = None
            for sheet_name in workbook.sheetnames:
                if "scopus" in sheet_name.lower() or "source" in sheet_name.lower():
                    sheet = workbook[sheet_name]
                    break

            if sheet is None:
                sheet = workbook.active

            if sheet is None:
                status_logger.error(
                    f"    {self.get_name()}: No valid sheet found in Excel file"
                )
                return []

            detail_logger.info(f"Reading sheet: {sheet.title}")

            # Read header row and find column indices
            rows_iter = iter(sheet.rows)
            header_row = next(rows_iter)
            headers = [cell.value for cell in header_row]
            col_indices = self._find_column_indices(headers)

            # Validate required columns exist
            if "title" not in col_indices:
                status_logger.error(
                    f"    {self.get_name()}: Could not find 'Source Title' column in file"
                )
                return []

            detail_logger.info(f"Found columns: {list(col_indices.keys())}")

            journals = []
            active_count = 0
            inactive_count = 0
            quality_flagged_count = 0

            # Process data rows
            for row in rows_iter:
                # Extract values using helper method
                title = self._extract_cell_value(row, col_indices, "title")
                issn = self._extract_cell_value(row, col_indices, "issn")
                eissn = self._extract_cell_value(row, col_indices, "eissn")
                publisher = self._extract_cell_value(row, col_indices, "publisher")
                status = self._extract_cell_value(row, col_indices, "status")
                quality_flag = self._extract_cell_value(
                    row, col_indices, "quality_flag"
                )
                source_type = self._extract_cell_value(row, col_indices, "source_type")
                coverage = self._extract_cell_value(row, col_indices, "coverage")
                open_access = self._extract_cell_value(row, col_indices, "open_access")

                # Skip rows without title
                if not title:
                    continue

                title = str(title).strip()
                if not title or len(title) < 2:
                    continue

                # Track statistics
                status_str = str(status).strip() if status else ""
                is_active = status_str.lower() == "active"
                is_quality_flagged = bool(quality_flag and str(quality_flag).strip())

                if is_active:
                    active_count += 1
                else:
                    inactive_count += 1

                if is_quality_flagged:
                    quality_flagged_count += 1

                # Only include active journals (skip inactive ones)
                if not is_active:
                    continue

                # Validate and normalize ISSNs
                issn_str = str(issn) if issn else None
                eissn_str = str(eissn) if eissn else None
                issn = self._validate_and_normalize_issn(issn_str, title)
                eissn = self._validate_and_normalize_issn(eissn_str, title)

                # Create journal entry
                journal_entry = self._create_journal_entry(
                    title,
                    issn,
                    eissn,
                    publisher,
                    source_type,
                    coverage,
                    open_access,
                    is_quality_flagged,
                    quality_flag,
                )

                if journal_entry:
                    journals.append(journal_entry)

            workbook.close()

            status_logger.info(
                f"    {self.get_name()}: Processed {len(journals)} active journals "
                f"({inactive_count} inactive excluded, {quality_flagged_count} quality-flagged found)"
            )

            return journals

        except Exception as e:
            status_logger.error(
                f"    {self.get_name()}: Error loading journal list - {e}"
            )
            return []
