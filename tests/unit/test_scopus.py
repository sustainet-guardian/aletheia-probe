# SPDX-License-Identifier: MIT
"""Tests for Scopus source and backend."""

import tempfile
from pathlib import Path
from unittest.mock import AsyncMock, Mock, patch

import pytest
from openpyxl import Workbook

from aletheia_probe.backends.scopus import ScopusBackend
from aletheia_probe.models import BackendStatus, QueryInput
from aletheia_probe.updater.sources import ScopusSource


class TestScopusSource:
    """Test cases for ScopusSource."""

    def test_get_name(self):
        """Test get_name returns 'scopus'."""
        source = ScopusSource()
        assert source.get_name() == "scopus"

    def test_get_list_type(self):
        """Test get_list_type returns 'legitimate'."""
        source = ScopusSource()
        assert source.get_list_type() == "legitimate"

    def test_should_update_no_file(self):
        """Test should_update returns False when file doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source = ScopusSource(data_dir=Path(tmpdir))
            assert source.should_update() is False

    def test_should_update_with_file(self):
        """Test should_update returns True when file exists and not updated."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            # Create a mock Scopus file
            test_file = data_dir / "ext_list_October_2024.xlsx"
            wb = Workbook()
            wb.save(test_file)
            wb.close()

            source = ScopusSource(data_dir=data_dir)

            with patch(
                "aletheia_probe.updater.sources.scopus.get_cache_manager"
            ) as mock_get_cache_manager:
                mock_cache_manager = Mock()
                mock_cache_manager.get_source_last_updated.return_value = None
                mock_get_cache_manager.return_value = mock_cache_manager
                assert source.should_update() is True

    def test_find_scopus_file_not_found(self):
        """Test _find_scopus_file when directory doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir) / "nonexistent"
            source = ScopusSource(data_dir=data_dir)
            assert source._find_scopus_file() is False

    def test_find_scopus_file_no_matching_files(self):
        """Test _find_scopus_file when no Excel files exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            source = ScopusSource(data_dir=data_dir)
            assert source._find_scopus_file() is False

    def test_find_scopus_file_success(self):
        """Test _find_scopus_file finds the file successfully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            # Create a mock Scopus file
            test_file = data_dir / "ext_list_October_2024.xlsx"
            wb = Workbook()
            wb.save(test_file)
            wb.close()

            source = ScopusSource(data_dir=data_dir)
            assert source._find_scopus_file() is True
            assert source.file_path == test_file

    def test_find_scopus_file_multiple_files(self):
        """Test _find_scopus_file selects most recent file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)

            # Create multiple files
            file1 = data_dir / "ext_list_January_2024.xlsx"
            file2 = data_dir / "ext_list_October_2024.xlsx"

            wb = Workbook()
            wb.save(file1)
            wb.close()

            wb = Workbook()
            wb.save(file2)
            wb.close()

            source = ScopusSource(data_dir=data_dir)
            assert source._find_scopus_file() is True
            # Should select the most recently modified file
            assert source.file_path in [file1, file2]

    @pytest.mark.asyncio
    async def test_fetch_data_no_file(self):
        """Test fetch_data returns empty list when file not found."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source = ScopusSource(data_dir=Path(tmpdir))
            data = await source.fetch_data()
            assert data == []

    @pytest.mark.asyncio
    async def test_fetch_data_success(self):
        """Test fetch_data successfully parses Scopus Excel file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            test_file = data_dir / "ext_list_October_2024.xlsx"

            # Create a mock Scopus Excel file with proper structure
            wb = Workbook()
            ws = wb.active
            assert ws is not None  # Type assertion for mypy
            ws.title = "Scopus Sources Oct. 2024"

            # Add headers
            headers = [
                "Source Title",
                "ISSN",
                "EISSN",
                "Publisher",
                "Active or Inactive",
                "Titles Discontinued by Scopus Due to Quality Issues",
                "Source Type",
                "Coverage",
                "Open Access Status",
            ]
            ws.append(headers)

            # Add test data - active journal
            ws.append(
                [
                    "Test Journal of Science",
                    "1234-5678",
                    "2049-3630",
                    "Test Publisher",
                    "Active",
                    "",
                    "Journal",
                    "2020-2024",
                    "Open Access",
                ]
            )

            # Add test data - inactive journal (should be filtered out)
            ws.append(
                [
                    "Inactive Journal",
                    "9999-9999",
                    "",
                    "Old Publisher",
                    "Inactive",
                    "",
                    "Journal",
                    "2010-2015",
                    "",
                ]
            )

            # Add test data - quality flagged journal
            ws.append(
                [
                    "Flagged Journal",
                    "1545-5971",
                    "",
                    "Questionable Publisher",
                    "Active",
                    "Poor editorial practices",
                    "Journal",
                    "2018-2020",
                    "",
                ]
            )

            wb.save(test_file)
            wb.close()

            source = ScopusSource(data_dir=data_dir)

            with patch(
                "aletheia_probe.normalizer.input_normalizer.normalize"
            ) as mock_normalize:
                # Mock the normalizer to return different normalized names
                def normalize_side_effect(name):
                    return Mock(normalized_name=name.lower())

                mock_normalize.side_effect = normalize_side_effect

                data = await source.fetch_data()

                # Should have 2 active journals (inactive one excluded)
                assert len(data) == 2

                # Check first journal
                assert data[0]["journal_name"] == "Test Journal of Science"
                assert data[0]["issn"] == "1234-5678"
                assert data[0]["eissn"] == "2049-3630"
                assert data[0]["publisher"] == "Test Publisher"
                assert "quality_flagged" not in data[0]["metadata"]

                # Check quality-flagged journal
                flagged = [j for j in data if j["journal_name"] == "Flagged Journal"][0]
                assert flagged["metadata"]["quality_flagged"] is True
                assert (
                    "Poor editorial practices"
                    in flagged["metadata"]["quality_flag_reason"]
                )

    @pytest.mark.asyncio
    async def test_fetch_data_missing_columns(self):
        """Test fetch_data handles missing required columns gracefully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            test_file = data_dir / "ext_list_October_2024.xlsx"

            # Create a file without required columns
            wb = Workbook()
            ws = wb.active
            assert ws is not None  # Type assertion for mypy
            ws.append(["Random Column", "Another Column"])
            ws.append(["Data 1", "Data 2"])

            wb.save(test_file)
            wb.close()

            source = ScopusSource(data_dir=data_dir)
            data = await source.fetch_data()

            # Should return empty list due to missing required columns
            assert data == []

    @pytest.mark.asyncio
    async def test_fetch_data_invalid_issn_format(self):
        """Test fetch_data handles invalid ISSN formats."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            test_file = data_dir / "ext_list_October_2024.xlsx"

            wb = Workbook()
            ws = wb.active
            assert ws is not None  # Type assertion for mypy
            ws.title = "Scopus Sources Oct. 2024"

            headers = ["Source Title", "ISSN", "Active or Inactive"]
            ws.append(headers)

            # Journal with invalid ISSN
            ws.append(["Test Journal", "INVALID", "Active"])

            wb.save(test_file)
            wb.close()

            source = ScopusSource(data_dir=data_dir)

            with patch(
                "aletheia_probe.normalizer.input_normalizer.normalize"
            ) as mock_normalize:
                mock_normalize.return_value = Mock(normalized_name="test journal")
                data = await source.fetch_data()

                # Should process the journal but with None ISSN
                assert len(data) == 1
                assert data[0]["issn"] is None


class TestScopusBackend:
    """Test cases for ScopusBackend."""

    def test_get_name(self):
        """Test get_name returns 'scopus'."""
        backend = ScopusBackend()
        assert backend.get_name() == "scopus"

    def test_get_description(self):
        """Test get_description returns expected string."""
        backend = ScopusBackend()
        description = backend.get_description()
        assert "scopus" in description.lower()
        assert "legitimate" in description.lower()

    def test_backend_configuration(self):
        """Test backend is configured correctly."""
        backend = ScopusBackend()
        assert backend.source_name == "scopus"
        assert backend.list_type == "legitimate"
        assert backend.cache_ttl_hours == 24 * 30  # Monthly cache

    @pytest.mark.asyncio
    async def test_query_journal_found(self):
        """Test querying a journal that exists in Scopus cache."""
        backend = ScopusBackend()
        query_input = QueryInput(
            raw_input="Test Journal",
            normalized_name="test journal",
            identifiers={"issn": "1234-5678"},
        )

        mock_results = [
            {
                "journal_name": "Test Journal",
                "normalized_name": "test journal",
                "issn": "1234-5678",
                "publisher": "Test Publisher",
                "metadata": {"source_type": "Journal"},
            }
        ]

        with patch(
            "aletheia_probe.backends.base.get_cache_manager"
        ) as mock_get_cache_manager:
            mock_cache = Mock()
            mock_cache.search_journals.return_value = mock_results
            mock_cache.search_journals_by_name.return_value = mock_results
            mock_get_cache_manager.return_value = mock_cache
            result = await backend.query(query_input)

            assert result.status == BackendStatus.FOUND
            assert result.assessment == "legitimate"
            assert result.confidence > 0.9  # ISSN match gives high confidence

    @pytest.mark.asyncio
    async def test_query_journal_not_found(self):
        """Test querying a journal that doesn't exist in Scopus cache."""
        backend = ScopusBackend()
        query_input = QueryInput(
            raw_input="Unknown Journal", normalized_name="unknown journal"
        )

        with patch(
            "aletheia_probe.backends.base.get_cache_manager"
        ) as mock_get_cache_manager:
            mock_cache = Mock()
            mock_cache.search_journals.return_value = []
            mock_cache.search_journals_by_name.return_value = []
            mock_get_cache_manager.return_value = mock_cache
            result = await backend.query(query_input)

            assert result.status == BackendStatus.NOT_FOUND
            assert result.assessment is None
            assert result.confidence == 0.0

    @pytest.mark.asyncio
    async def test_query_with_quality_flagged_journal(self):
        """Test querying a quality-flagged journal still returns as legitimate."""
        backend = ScopusBackend()
        query_input = QueryInput(
            raw_input="Flagged Journal",
            normalized_name="flagged journal",
            identifiers={"issn": "8888-8888"},
        )

        mock_results = [
            {
                "journal_name": "Flagged Journal",
                "normalized_name": "flagged journal",
                "issn": "8888-8888",
                "metadata": {
                    "quality_flagged": True,
                    "quality_flag_reason": "Poor practices",
                },
            }
        ]

        with patch(
            "aletheia_probe.backends.base.get_cache_manager"
        ) as mock_get_cache_manager:
            mock_cache = Mock()
            mock_cache.search_journals.return_value = mock_results
            mock_cache.search_journals_by_name.return_value = mock_results
            mock_get_cache_manager.return_value = mock_cache
            result = await backend.query(query_input)

            # Still returns as legitimate (indexed in Scopus)
            # but the quality flag is in metadata
            assert result.status == BackendStatus.FOUND
            assert result.assessment == "legitimate"
            assert result.data["source_data"]["metadata"]["quality_flagged"] is True
